# app.py
# -*- coding: utf-8 -*-
import os, io, json, sqlite3, hashlib, secrets, re
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from dataclasses import dataclass
from typing import Dict, Optional, Tuple, List

import streamlit as st
import pandas as pd
import pdfplumber
import altair as alt

# =================== CONFIG GERAL ===================
st.set_page_config(page_title="Wofair • Loterias", page_icon="🔐", layout="wide")

DATA_DIR = "data"
DB_PATH = os.path.join(DATA_DIR, "app.db")
os.makedirs(DATA_DIR, exist_ok=True)

def q2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

# =================== BANCO (SQLite) ===================
@st.cache_resource
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    init_db(conn)
    return conn

def init_db(conn):
    conn.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        salt TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    """)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        period_start TEXT,
        period_end TEXT,
        file_name TEXT,
        total_esperado REAL NOT NULL,
        items_json TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_results_user_period
        ON results(user_id, period_start, period_end);
    """)
    conn.commit()

def hash_password(password: str, salt: Optional[bytes]=None) -> Tuple[str,str]:
    if salt is None:
        salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return salt.hex(), dk.hex()

def verify_password(password: str, salt_hex: str, hash_hex: str) -> bool:
    salt = bytes.fromhex(salt_hex)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return secrets.compare_digest(dk.hex(), hash_hex)

def create_user(conn, email: str, password: str) -> Tuple[bool, str]:
    try:
        salt_hex, hash_hex = hash_password(password)
        conn.execute(
            "INSERT INTO users (email, salt, password_hash, created_at) VALUES (?, ?, ?, ?)",
            (email.lower(), salt_hex, hash_hex, datetime.utcnow().isoformat())
        )
        conn.commit()
        return True, "Conta criada com sucesso!"
    except sqlite3.IntegrityError:
        return False, "Email já cadastrado."
    except Exception as e:
        return False, f"Erro ao criar conta: {e}"

def get_user_by_email(conn, email: str):
    cur = conn.execute("SELECT id, email, salt, password_hash, created_at FROM users WHERE email = ?", (email.lower(),))
    row = cur.fetchone()
    return row  # tuple or None

def save_result(conn, user_id: int, period_start: Optional[str], period_end: Optional[str], file_name: str,
                total_esperado: float, items_df: pd.DataFrame):
    payload = items_df.to_dict(orient="records")
    conn.execute("""
    INSERT INTO results (user_id, created_at, period_start, period_end, file_name, total_esperado, items_json)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, datetime.utcnow().isoformat(), period_start, period_end, file_name, float(total_esperado), json.dumps(payload, ensure_ascii=False)))
    conn.commit()

def list_results(conn, user_id: int, limit: int = 50) -> pd.DataFrame:
    cur = conn.execute("""
      SELECT id, created_at, period_start, period_end, file_name, total_esperado
      FROM results
      WHERE user_id = ?
      ORDER BY id DESC
      LIMIT ?
    """, (user_id, limit))
    rows = cur.fetchall()
    return pd.DataFrame(rows, columns=["id","created_at","period_start","period_end","file_name","total_esperado"])

def load_result_items(conn, user_id: int, result_id: int) -> Optional[pd.DataFrame]:
    cur = conn.execute("SELECT items_json FROM results WHERE id=? AND user_id=?", (result_id, user_id))
    row = cur.fetchone()
    if not row: return None
    data = json.loads(row[0])
    return pd.DataFrame(data)

def upsert_result_by_start(conn, user_id: int,
                           period_start_br: Optional[str], period_end_br: Optional[str],
                           file_name: str, total_esperado: float, items_df: pd.DataFrame) -> tuple[str, int]:
    """
    Atualiza se já existir um resultado do MESMO usuário com o MESMO period_start (comparado por data).
    Caso contrário, insere. Datas são salvas em 'DD/MM/AAAA'.
    """
    payload = items_df.to_dict(orient="records")
    in_start_iso = _pdfdate_to_iso_str(period_start_br)

    # Se não tem início legível -> nunca atualiza, sempre insere
    if not in_start_iso:
        conn.execute("""
            INSERT INTO results (user_id, created_at, period_start, period_end, file_name, total_esperado, items_json)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user_id, datetime.utcnow().isoformat(), period_start_br, period_end_br, file_name, float(total_esperado),
              json.dumps(payload, ensure_ascii=False)))
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return ("inserted", rid)

    # Procura por início igual (mesmo dia)
    cur = conn.execute("SELECT id, period_start FROM results WHERE user_id = ?", (user_id,))
    match_id = None
    for rid, db_start in cur.fetchall():
        if _pdfdate_to_iso_str(db_start) == in_start_iso:
            match_id = rid
            break

    now_iso = datetime.utcnow().isoformat()
    if match_id is not None:
        conn.execute("""
            UPDATE results
               SET created_at = ?, period_start = ?, period_end = ?, file_name = ?, total_esperado = ?, items_json = ?
             WHERE id = ? AND user_id = ?
        """, (now_iso, period_start_br, period_end_br, file_name, float(total_esperado),
              json.dumps(payload, ensure_ascii=False), match_id, user_id))
        conn.commit()
        return ("updated", match_id)

    conn.execute("""
        INSERT INTO results (user_id, created_at, period_start, period_end, file_name, total_esperado, items_json)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, now_iso, period_start_br, period_end_br, file_name, float(total_esperado),
          json.dumps(payload, ensure_ascii=False)))
    conn.commit()
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return ("inserted", rid)


def delete_results(conn, user_id: int) -> int:
    cur = conn.execute("DELETE FROM results WHERE user_id = ?", (user_id,))
    conn.commit()
    return cur.rowcount

def delete_user_and_results(conn, user_id: int) -> tuple[int, int]:
    # apaga resultados primeiro, depois a conta
    cur_res = conn.execute("DELETE FROM results WHERE user_id = ?", (user_id,))
    cur_usr = conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    return cur_res.rowcount, cur_usr.rowcount

def dedupe_by_start(conn, user_id: int) -> int:
    # pega o id mais novo por period_start
    keep = dict(conn.execute("""
        SELECT period_start, MAX(id) AS keep_id
          FROM results
         WHERE user_id = ?
         GROUP BY period_start
    """, (user_id,)).fetchall())
    # seleciona todos e marca os que não são keepers
    rows = conn.execute("SELECT id, period_start FROM results WHERE user_id = ?", (user_id,)).fetchall()
    to_delete = [rid for (rid, ps) in rows if ps in keep and rid != keep[ps]]
    if to_delete:
        conn.executemany("DELETE FROM results WHERE id = ?", [(rid,) for rid in to_delete])
        conn.commit()
    return len(to_delete)

# =================== FUNÇÕES PARA DATAS ===================
# ---- Datas do PDF/DB -> date ----
_PT_MONTH = {"JAN":1,"FEV":2,"MAR":3,"ABR":4,"MAI":5,"JUN":6,"JUL":7,"AGO":8,"SET":9,"OUT":10,"NOV":11,"DEZ":12}

def _parse_pdf_date(s: Optional[str]) -> Optional[date]:
    if not s: return None
    s = s.strip().upper()
    # 08/AGO/25
    m = re.match(r"^(\d{2})/([A-Z]{3})/(\d{2})$", s)
    if m:
        d, mon, yy = int(m.group(1)), m.group(2), int(m.group(3))
        yyyy = 2000 + yy if yy < 70 else 1900 + yy
        return date(yyyy, _PT_MONTH.get(mon, 1), d)
    # 02.07.2025
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # 11/08/2025
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None

def _row_effective_date(row: dict) -> Optional[date]:
    # 1) usa fim do período, 2) início, 3) created_at
    for key in ("period_end","period_start"):
        d = _parse_pdf_date(row.get(key))
        if d: return d
    try:
        return datetime.fromisoformat(row.get("created_at","")).date()
    except Exception:
        return None

def _load_concat_items(conn, user_id: int, ids: List[int]) -> pd.DataFrame:
    frames = []
    for rid in ids:
        df_i = load_result_items(conn, user_id, int(rid))
        if df_i is not None and not df_i.empty:
            df_i = df_i.copy()
            df_i["result_id"] = rid  # <- importante p/ ligar com a data do resultado
            frames.append(df_i)
    cols = ["Seção","Tipo","Qtde","Valor Unitário","Valor Esperado","result_id"]
    return pd.concat(frames, ignore_index=True)[cols] if frames else pd.DataFrame(columns=cols)


def date_input_br(label: str, *, value: date, min_value: date = None, max_value: date = None, key: str = None):
    """
    Date input com formato brasileiro. Em versões novas do Streamlit usa `format="DD/MM/YYYY"`;
    em versões antigas, mantém o date_input padrão, mas deixa o label explícito.
    """
    try:
        # Streamlit >= 1.29 (aprox.) – tem suporte a 'format'
        return st.date_input(label, value=value, min_value=min_value, max_value=max_value, key=key, format="DD/MM/YYYY")
    except TypeError:
        # Streamlit mais antigo – sem 'format'
        return st.date_input(f"{label} (dd/mm/aaaa)", value=value, min_value=min_value, max_value=max_value, key=key)

def _pdfdate_to_iso_str(s: Optional[str]) -> Optional[str]:
    """Converte '01/AGO/25' (ou variações) para '2025-08-01'. Se não entender, retorna None."""
    d = _parse_pdf_date(s)
    return d.isoformat() if d else None

def _pdfdate_to_br_str(s: Optional[str]) -> Optional[str]:
    """Qualquer formato aceito pelo _parse_pdf_date -> 'DD/MM/AAAA'."""
    d = _parse_pdf_date(s)
    return d.strftime("%d/%m/%Y") if d else None

def _is_iso_yyyy_mm_dd(s: Optional[str]) -> bool:
    return bool(s and re.fullmatch(r"\d{4}-\d{2}-\d{2}", s))

# =================== FUNÇÕES PARA ABAS ===================
def render_uploads(conn, user):
    st.markdown("### 📥 Uploads ")
    files = st.file_uploader(
        "Envie PDF(s) do relatório diário",
        type=["pdf"], accept_multiple_files=True, key="uploader_all"
    )

    if not files:
        st.caption("Selecione 1 PDF para processar imediatamente, ou vários para processamento em lote.")
        return

    # === 1 arquivo: fluxo imediato no topo ===
    if len(files) == 1:
        uploaded = files[0]
        try:
            df, periodo, total = _process_single_pdf(uploaded)
        except Exception as e:
            st.error(f"Falha ao processar PDF: {e}")
            return

        total_esperado = total
        faltando_unit = int(df.loc[df["Tipo"] != "*TELESENA", "Valor Unitário"].isna().sum())
        c1, c2 = st.columns(2)
        fmt = lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        c1.metric("Total Esperado (todos os serviços)", fmt(total_esperado))
        c2.metric("Tipos sem valor unitário definido", str(faltando_unit))

        st.markdown("#### 🧾 Itens detectados (apenas remunerados)")
        st.dataframe(df[["Seção","Tipo","Qtde","Valor PDF","Valor Unitário","Valor Esperado"]],
                     use_container_width=True, hide_index=True)

        st.markdown("#### 📚 Resumos")
        colA, colB = st.columns(2)
        resumo_tipo = (
            df.groupby("Tipo", as_index=False)
              .agg(Qtde=("Qtde","sum"), Valor_Esperado=("Valor Esperado","sum"))
              .sort_values(["Valor_Esperado","Qtde"], ascending=[False, False])
        )
        colA.write("**Por Tipo**"); colA.dataframe(resumo_tipo, use_container_width=True, hide_index=True)
        resumo_secao = (
            df.groupby("Seção", as_index=False)
              .agg(Qtde=("Qtde","sum"), Valor_Esperado=("Valor Esperado","sum"))
              .sort_values("Valor_Esperado", ascending=False)
        )
        colB.write("**Por Seção**"); colB.dataframe(resumo_secao, use_container_width=True, hide_index=True)

        per_ini = _pdfdate_to_br_str(periodo[0]) if (periodo and periodo[0]) else None
        per_fim = _pdfdate_to_br_str(periodo[1]) if (periodo and len(periodo) > 1 and periodo[1]) else None

        if st.button("💾 Salvar este resultado", key="save_single"):
            status, rid = upsert_result_by_start(conn, user["id"], per_ini, per_fim, uploaded.name, float(total_esperado), df)
            if status == "updated":
                st.success(f"Atualizado para {per_ini or '-'} — {per_fim or '-'} (id {rid}).")
            else:
                st.success(f"Inserido para {per_ini or '-'} — {per_fim or '-'} (id {rid}).")
            _safe_rerun()

    # === 2+ arquivos: opções + botão para rodar ===
    else:
        st.info(f"{len(files)} arquivo(s) selecionado(s).")
        col1, col2 = st.columns(2)
        with col1:
            overwrite = st.checkbox("Atualizar se já existir mesmo início (upsert)", value=True, key="batch_overwrite")
        with col2:
            dry_run   = st.checkbox("Somente simular (não salvar no banco)", value=False, key="batch_dry_run")

        if st.button("🚀 Processar lote", use_container_width=True, key="btn_processar_lote"):
            # Use a versão parametrizada do process_batch (overwrite/dry_run)
            process_batch(conn, user["id"], files, overwrite=overwrite, dry_run=dry_run)


def render_historico(conn, user):
    st.markdown("### 📜 Meu histórico")
    df_hist = list_results(conn, user["id"])
    if df_hist.empty:
        st.caption("Nenhum resultado salvo ainda.")
        return

    df_hist = df_hist.copy()
    df_hist["ps_date"] = df_hist["period_start"].apply(_parse_pdf_date)
    df_hist = df_hist.sort_values("ps_date", ascending=False)

    st.dataframe(df_hist[["id","period_start","period_end","file_name","total_esperado"]],
                 use_container_width=True, hide_index=True)

    datas_opcoes = df_hist["period_start"].dropna().unique().tolist()
    sel_data = st.selectbox("Abrir resultado por data (início):", datas_opcoes,
                            index=0 if datas_opcoes else None, key="open_by_date")
    if st.button("Abrir", key="abrir_por_data"):
        rid = int(df_hist.loc[df_hist["period_start"] == sel_data, "id"].iloc[0])
        df_saved = load_result_items(conn, user["id"], rid)
        if df_saved is None or df_saved.empty:
            st.error("Resultado não encontrado para essa data.")
        else:
            st.dataframe(df_saved, use_container_width=True, hide_index=True)


def render_relatorio_periodo(conn, user):
    st.markdown("### 📆 Relatório por Período")
    df_hist = list_results(conn, user["id"])
    if df_hist.empty:
        st.caption("Você ainda não salvou resultados. Gere e salve pelo menos um para usar o filtro por período.")
        return

    df_hist = df_hist.copy()
    df_hist["eff_date"] = df_hist.apply(lambda r: _row_effective_date(r), axis=1)
    dmin = min([d for d in df_hist["eff_date"] if d is not None], default=date.today())
    dmax = max([d for d in df_hist["eff_date"] if d is not None], default=date.today())

    colf1, colf2, colf3 = st.columns([1,1,1])
    with colf1:
        start_date = date_input_br("Início", value=dmin, key="inicio_br")
    with colf2:
        end_date   = date_input_br("Fim", value=dmax, key="fim_br")
    with colf3:
        ate_hoje   = st.toggle("Até hoje", value=False, key="ate_hoje")
    if ate_hoje:
        end_date = date.today()

    mask = df_hist["eff_date"].apply(lambda d: d is not None and (start_date <= d <= end_date))
    df_sel = df_hist[mask].copy()
    st.caption(f"{len(df_sel)} resultado(s) no período {start_date.strftime('%d/%m/%Y')} — {end_date.strftime('%d/%m/%Y')}.")

    if df_sel.empty:
        st.info("Nenhum resultado nesse intervalo.")
        return

    # opcional: só o mais recente por início
    use_latest = st.checkbox("Usar apenas o resultado mais recente por data de início", value=True, key="use_latest_start")
    if use_latest:
        df_sel = (
            df_sel.assign(ps_key=df_sel["period_start"].fillna(""))
                  .sort_values("id", ascending=False)
                  .drop_duplicates(subset=["ps_key"], keep="first")
                  .drop(columns=["ps_key"])
                  .sort_values("eff_date")
        )

    ids = df_sel["id"].tolist()
    df_all = _load_concat_items(conn, user["id"], ids)
    for col in ["Qtde","Valor Unitário","Valor Esperado","Valor PDF"]:
        if col in df_all.columns:
            df_all[col] = pd.to_numeric(df_all[col], errors="coerce")

    tot_esperado = float(df_all["Valor Esperado"].fillna(0).sum()) if "Valor Esperado" in df_all.columns else 0.0
    num_docs = len(ids)
    c1, c2 = st.columns(2)
    c1.metric("Total Esperado no Período", f"R$ {tot_esperado:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    c2.metric("Resultados considerados", f"{num_docs}")

    st.markdown("#### 🔗 Itens consolidados")
    st.dataframe(df_all, use_container_width=True, hide_index=True)

    st.markdown("#### 📚 Consolidação por Tipo")
    by_tipo = (
        df_all.groupby("Tipo", as_index=False)
              .agg(Qtde=("Qtde","sum"), Valor_Esperado=("Valor Esperado","sum"))
              .sort_values(["Valor_Esperado","Qtde"], ascending=[False, False])
    )
    st.dataframe(by_tipo, use_container_width=True, hide_index=True)

    st.markdown("#### 📦 Consolidação por Seção")
    by_secao = (
        df_all.groupby("Seção", as_index=False)
              .agg(Qtde=("Qtde","sum"), Valor_Esperado=("Valor Esperado","sum"))
              .sort_values("Valor_Esperado", ascending=False)
    )
    st.dataframe(by_secao, use_container_width=True, hide_index=True)

    # === Gráficos Altair ===
    map_dates = df_sel.set_index("id")["eff_date"].to_dict()
    df_all["eff_date"] = df_all["result_id"].map(map_dates)

    serie_dia = (
        df_all.dropna(subset=["eff_date"])
              .groupby("eff_date", as_index=False)["Valor Esperado"]
              .sum()
              .rename(columns={"Valor Esperado": "Total"})
              .sort_values("eff_date")
    )

    st.markdown("#### 📈 Evolução diária no período")
    if not serie_dia.empty:
        c_line = (
            alt.Chart(serie_dia)
               .mark_line(point=True)
               .encode(
                   x=alt.X("eff_date:T", title="Data", axis=alt.Axis(format="%d/%m/%Y")),
                   y=alt.Y("Total:Q", title="Total esperado (R$)", stack=None),
                   tooltip=[alt.Tooltip("eff_date:T", title="Data", format="%d/%m/%Y"),
                            alt.Tooltip("Total:Q", title="Total (R$)", format=",.2f")]
               )
               .properties(height=320)
               .interactive()
        )
        st.altair_chart(c_line, use_container_width=True)

    st.markdown("#### 🥇 Top 10 • Tipos no período")
    if not by_tipo.empty:
        top10 = by_tipo.head(10).copy()
        c_bar_tipos = (
            alt.Chart(top10)
               .mark_bar()
               .encode(
                   x=alt.X("Valor_Esperado:Q", title="R$ no período", axis=alt.Axis(format=",.2f")),
                   y=alt.Y("Tipo:N", sort="-x", title=None),
                   tooltip=[alt.Tooltip("Tipo:N", title="Tipo"),
                            alt.Tooltip("Qtde:Q", title="Qtde"),
                            alt.Tooltip("Valor_Esperado:Q", title="R$ no período", format=",.2f")]
               )
               .properties(height=max(220, 22*len(top10)))
        )
        st.altair_chart(c_bar_tipos, use_container_width=True)

    st.markdown("#### 🧩 Evolução por Seção")
    df_sec_dia = (
        df_all.dropna(subset=["eff_date"])
              .groupby(["eff_date","Seção"], as_index=False)["Valor Esperado"]
              .sum()
              .rename(columns={"Valor Esperado":"Total"})
              .sort_values(["eff_date","Seção"])
    )
    if not df_sec_dia.empty:
        c_area = (
            alt.Chart(df_sec_dia)
               .mark_area()
               .encode(
                   x=alt.X("eff_date:T", title="Data", axis=alt.Axis(format="%d/%m/%Y")),
                   y=alt.Y("Total:Q", title="Total esperado (R$)", stack="zero"),
                   color=alt.Color("Seção:N", title="Seção"),
                   tooltip=[alt.Tooltip("eff_date:T", title="Data", format="%d/%m/%Y"),
                            alt.Tooltip("Seção:N", title="Seção"),
                            alt.Tooltip("Total:Q", title="R$ no dia", format=",.2f")]
               )
               .properties(height=340)
               .interactive()
        )
        st.altair_chart(c_area, use_container_width=True)

    # === Export Excel consolidado ===
    st.markdown("#### ⬇️ Exportar consolidação")
    buffer = io.BytesIO()
    try:
        from openpyxl.utils import get_column_letter
        engine = "openpyxl"; OPENPYXL_OK = True
    except Exception:
        engine = "xlsxwriter"; OPENPYXL_OK = False

    with pd.ExcelWriter(buffer, engine=engine) as writer:
        df_sel.to_excel(writer, index=False, sheet_name="Resultados (metadados)")
        df_all.to_excel(writer, index=False, sheet_name="Itens (consolidado)")
        by_tipo.to_excel(writer, index=False, sheet_name="Por Tipo")
        by_secao.to_excel(writer, index=False, sheet_name="Por Seção")

        if OPENPYXL_OK:
            ws = writer.sheets["Itens (consolidado)"]
            for i, w in enumerate([12, 36, 10, 14, 16], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            for sheet_name, df_ref in [("Itens (consolidado)", df_all), ("Por Tipo", by_tipo), ("Por Seção", by_secao)]:
                ws2 = writer.sheets[sheet_name]
                for col_name in ("Valor Esperado","Valor Unitário","Valor PDF"):
                    if col_name in df_ref.columns:
                        cidx = df_ref.columns.get_loc(col_name) + 1
                        col = get_column_letter(cidx)
                        for r in range(2, ws2.max_row + 1):
                            ws2[f"{col}{r}"].number_format = 'R$ #,##0.00'
        else:
            book = writer.book
            currency = book.add_format({'num_format': 'R$ #,##0.00'})
            if not df_all.empty:
                sh = writer.sheets["Itens (consolidado)"]
                sh.set_column(0,0,12); sh.set_column(1,1,36); sh.set_column(2,2,10)
                for name in ("Valor Unitário","Valor Esperado","Valor PDF"):
                    if name in df_all.columns:
                        idx = df_all.columns.get_loc(name)
                        sh.set_column(idx, idx, 16 if name=="Valor Esperado" else 14, currency)

    st.download_button(
        "Baixar Excel consolidado (.xlsx)",
        buffer.getvalue(),
        "consolidado_periodo.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


# =================== AUTENTICAÇÃO (UI) ===================
def auth_ui(conn):
    st.sidebar.header("🔐 Acesso")
    mode = st.sidebar.radio("Escolha:", ["Entrar", "Criar conta"])
    if mode == "Entrar":
        with st.sidebar.form("login_form", clear_on_submit=False):
            email = st.text_input("Email")
            password = st.text_input("Senha", type="password")
            ok = st.form_submit_button("Entrar")
        if ok:
            user = get_user_by_email(conn, email)
            if not user:
                st.sidebar.error("Usuário não encontrado.")
            else:
                uid, uemail, salt_hex, hash_hex, _ = user
                if verify_password(password, salt_hex, hash_hex):
                    st.session_state["user"] = {"id": uid, "email": uemail}
                    st.sidebar.success("Login ok!")
                else:
                    st.sidebar.error("Senha inválida.")
    else:  # Criar conta
        with st.sidebar.form("signup_form", clear_on_submit=False):
            email = st.text_input("Email")
            pw1 = st.text_input("Senha", type="password")
            pw2 = st.text_input("Confirmar senha", type="password")
            ok = st.form_submit_button("Criar conta")
        if ok:
            if not email or not pw1:
                st.sidebar.error("Preencha email e senha.")
            elif pw1 != pw2:
                st.sidebar.error("As senhas não conferem.")
            else:
                okc, msg = create_user(conn, email, pw1)
                if okc:
                    st.sidebar.success(msg)
                else:
                    st.sidebar.error(msg)

    if "user" in st.session_state:
        u = st.session_state["user"]
        st.sidebar.info(f"Conectado: {u['email']}")
        if st.sidebar.button("Sair"):
            st.session_state.pop("user", None)
            _safe_rerun()

# =================== PARSER + CÁLCULO (seu fluxo) ===================
# Dicionário fixo (o mesmo que você aprovou)
UNITARIOS: Dict[str, float] = {
  "*SAQUE UL": 1.21,
  "BIOMETRIA SAQUE U": 1.144,
  "BIOM SQ CT DIGIT": 0.71,
  "BIOMETRIA SQ POU": 0.68,
  "*PAG INSS UL": 0.71,
  "*PAG INSS UL SOC": 0.71,
  "*SAQUE POUPANCA F": 0.68,
  "SAQUE EMERG S/CA": 0.78,
  "*SAQUE CT SOCIAL": 0.74,
  "*SQ CTA AUX BRASI": 0.78,
  "2-BENEF SOCIAIS": 0.89,
  "AGIBANK": 0.90,
  "*PAG FGTS": 1.28,
  "*DEPOSITO UL": 1.24,
  "2- CARTOES": 0.99,
  "*CAESB": 0.96,
  "*COPASA": 0.96,
  "*EMBASA": 0.96,
  "*SANEAGO": 0.96,
  "*C.AGUASLI": 0.96,
  "NEOENERGIA": 0.96,
  "CLIENT CO": 0.96,
  "OI": 0.96,
  "*VIVO SE": 0.96,
  "*TELEGOC": 0.96,
  "*TELEBC": 0.96,
  "*VIVO MG": 0.96,
  "*TELESP CL": 0.96,
  "*VIVOFIXO": 0.96,
  "*TIM": 0.96,
  "CLARO SA": 0.96,
  "*CLAROTV": 0.96,
  "*SKY": 0.96,
  "2-GPS S/ BARRA UL": 0.86,
  "*VIVO DF": 1.04,
  "*VIVO FIXO": 1.04,
  "*TIMCEL": 1.04,
  "*CLARODDDS": 1.04,
  "*DIRECTVSK": 1.04,
  "2-H AFINZ": 1.04,
  "SHOW DA FE": 1.04,
  "MUN CAB GRANDE": 0.96,
  "2-SEFAZ/GO-DARE": 0.96,
  "2-SEFAZ/DF": 0.96,
  "2-IPVA MG": 0.96,
  "2-DETRAN GO": 0.96,
  "GPS COM BARRA": 0.86,
  "2-SIMPLESNACION": 0.86,
  "DARF NUMERADO": 0.86,
  "DAE": 0.86,
  "2-DETRAN-DF": 0.96,
  "2-DER": 0.96,
  "*CLARO CO": 1.445,
  "*TCODFCE": 0.61857,
  "*TIM-GSM": 0.48,
  "*VIVO PI": 0.54,
  "PEC COBCAIXA": 0.94,
  "* NPC BOLETOS CAI": 0.94,
  "* NPC BOLETOS OUT": 1.24,
  "*PIX SAQUE": 0.74,
  "*PIX SAQUE CAIXA": 0.74,
  "*SALDO": 0.21,
  "*ATIVA CAIXA TEM": 1.02,
  "SAQUE EMERG S/CAR": 0.78,
  "*PIX PAGAMENTO CA": 0.00,
  "MEIO PAG C CREDIT": 0.00,
  "MEIO PAG C DEBITO": 0.00,
  "*PIX PAGAMENTO": 0.00,
  "*GERACAO QRCODE P": 0.00,
  "*TELESENA": 0.00,
  "*CAGECE": 0.96,
  "*CEMIG DIS": 0.96,
  "2-DAE SEF MG": 0.96,
  "2-PMDEBOAVISTA": 0.96,
  "*NBT": 0.96,
  "AGUASPIAUI": 0.96,
  "*SAAECABEC": 0.96,
  "*CLARO BSE": 0.99,
  "*PAG INSS UL SOCI": 0.71,
  "*CLARO TV": 1.04,
  "DEMAE": 0.96,
  "2-MULTAS DER MG": 0.96,
  "*VIVO PIAU": 0.96,
  "*NET": 1.04,
  "*ARREC SUPER XCAP": 1.00,
  "2-AMBIENT PAGSP": 0.96,
  "*EQUATORIA": 0.96,
  "2-PLANAL": 0.96,
  "*BAU PARCE": 0.96,
  "*TELMSCEL": 0.96,
  "BIOMETRIA SQ POUP": 0.68,
  "2- PORTO": 0.96,
  "*VIVO BA": 0.96,
  "*PARAIBA": 0.96,
  "2-SEFAZ/GO-OUTR": 0.96,
  "*TELMTCEL": 0.96,
  "*HUGHES": 0.96,
  "2-SEC.FIN.": 0.96,
  "*TELROCELU": 0.96,
  "*VIVO CEAR": 0.96

}


SECTION_HEADERS = {"PAGAMENTOS", "RECEBIMENTOS", "NEGOCIAL", "SERVICOS", "SERVIÇOS"}
ITEM_RE_STRICT = re.compile(
    r"""^(?P<tipo>[\*\wÀ-ÖØ-öø-ÿ0-9/\-\.\s]+?)   # tipo
        [ \t]{2,}
        (?P<qtde>\d+)
        [ \t]{2,}
        (?P<valor>[\d\.,]+)
        \s*$""",
    re.VERBOSE
)

ITEM_RE_FALLBACK = re.compile(
    r"""^(?P<tipo>.+?)\s+(?P<qtde>\d+)\s+(?P<valor>[\d\.,]+)\s*$""",
    re.VERBOSE
)


@dataclass
class Item:
    secao: str
    tipo: str
    qtde: int
    valor: Optional[Decimal] = None

def _money_br_to_decimal(s: str) -> Optional[Decimal]:
    if not s:
        return None
    s = s.strip()
    try:
        # "1.234,56" -> "1234.56"
        return Decimal(s.replace(".", "").replace(",", "."))
    except Exception:
        return None

def _normalize_spaces(s: str) -> str:
    return (s.replace("\u00A0", " ")
             .replace("\u2007", " ")
             .replace("\u2002", " ")
             .replace("\u2003", " "))

def extract_pdf_text(uploaded_file) -> str:
    with pdfplumber.open(uploaded_file) as pdf:
        parts = []
        for p in pdf.pages:
            txt = p.extract_text(x_tolerance=1, y_tolerance=1) or ""
            parts.append(txt)
        return "\n".join(parts)

def parse_items(text: str) -> Tuple[List[Item], Optional[Tuple[str, Optional[str]]]]:
    text = _normalize_spaces(text).replace("SERVIÇOS", "SERVICOS")
    lines = [ln.rstrip() for ln in text.splitlines()]
    # período (se houver 1 ou 2 datas)
    mper = re.search(r"(\d{2}/[A-Z]{3}/\d{2})(?:\s*A\s*(\d{2}/[A-Z]{3}/\d{2}))?", text)
    periodo = None
    if mper:
        periodo = (mper.group(1), mper.group(2) if mper.lastindex and mper.group(2) else None)

    current = None
    items: List[Item] = []

    for raw_ln in lines:
        s = _normalize_spaces(raw_ln).strip()
        if not s:
            continue
        if s in SECTION_HEADERS:
            current = s
            continue
        if not current:
            continue
        if s.startswith("TIPO") or s.startswith("TOTAL"):
            if s.startswith("TOTAL"):
                current = None
            continue

        m1 = ITEM_RE_STRICT.match(s)
        if m1:
            tipo  = re.sub(r"\s{2,}", " ", m1.group("tipo")).strip()
            qtde  = int(m1.group("qtde"))
            valor = _money_br_to_decimal(m1.group("valor"))
            items.append(Item(current, tipo, qtde, valor))  
            continue

        m2 = ITEM_RE_FALLBACK.match(s)
        if m2:
            tipo  = re.sub(r"\s{2,}", " ", m2.group("tipo")).strip()
            qtde  = int(m2.group("qtde"))
            if not tipo.upper().startswith("TIPO"):
                valor = _money_br_to_decimal(m2.group("valor"))
                items.append(Item(current, tipo, qtde, valor)) 
            continue


    return items, periodo

def _calc_df_from_items(items: List[Item]) -> pd.DataFrame:
    """Monta o DataFrame final a partir dos Items parseados, com as regras de unitário, zero e TELESENA."""
    df_raw = pd.DataFrame(
        [{
            "Seção": it.secao,
            "Tipo": it.tipo,
            "Qtde": it.qtde,
            "Valor PDF": float(it.valor) if getattr(it, "valor", None) is not None else None,
        } for it in items],
        columns=["Seção","Tipo","Qtde","Valor PDF"]
    )
    if df_raw.empty:
        return df_raw
    
    # mapeia unitários
    def unit_dec(tipo: str):
        vu = UNITARIOS.get(tipo)
        return None if vu is None else Decimal(str(vu))
    df_raw["Valor Unitário"] = df_raw["Tipo"].map(unit_dec)

    # não excluir TELESENA mesmo se unitário = 0
    zero_tipos = {t for t, v in UNITARIOS.items() if Decimal(str(v)) == Decimal("0") and t != "*TELESENA"}
    df = df_raw[~df_raw["Tipo"].isin(zero_tipos)].copy()

    # cálculo
    def valor_esperado(tipo: str, qtde: int, vu: Optional[Decimal], valor_pdf: Optional[float]):
        if isinstance(tipo, str) and tipo.upper() == "*TELESENA":
            if valor_pdf is None or pd.isna(valor_pdf):
                return None
            return float(q2(Decimal(str(valor_pdf)) * Decimal("0.08")))  # 8% do valor do PDF
        if vu is None: return None
        return float(q2(Decimal(qtde) * vu))

    df["Valor Esperado"] = df.apply(
        lambda r: valor_esperado(r["Tipo"], r["Qtde"], r["Valor Unitário"], r.get("Valor PDF")),
        axis=1
    )
    df["Valor Unitário"] = df["Valor Unitário"].astype(float)
    return df

def _process_single_pdf(file) -> tuple[pd.DataFrame, Optional[Tuple[str, Optional[str]]], Decimal]:
    """Lê o PDF, extrai items, monta DF final e retorna (df, periodo, total_esperado). Lança exceção se falhar."""
    raw = extract_pdf_text(file)
    items, periodo = parse_items(raw)
    if not items:
        raise ValueError("PDF sem itens reconhecíveis.")
    df = _calc_df_from_items(items)
    total = q2(sum(Decimal(str(v)) for v in df["Valor Esperado"].dropna())) if df["Valor Esperado"].notna().any() else Decimal("0.00")
    return df, periodo, total

def process_batch(conn, user_id: int, files: List, *, overwrite: bool = True, dry_run: bool = False) -> None:
    """Processa e (opcionalmente) salva vários PDFs de uma vez, com barra de progresso e resumo."""
    st.markdown("## 📦 Processamento em lote — execução")
    rows = []
    prog = st.progress(0)
    status = st.empty()

    total_files = len(files)
    for i, f in enumerate(files, start=1):
        fname = getattr(f, "name", f"arquivo_{i}.pdf")
        try:
            df, periodo, total = _process_single_pdf(f)
            per_ini = _pdfdate_to_br_str(periodo[0]) if (periodo and periodo[0]) else None
            per_fim = _pdfdate_to_br_str(periodo[1]) if (periodo and len(periodo) > 1 and periodo[1]) else None

            if not dry_run:
                if overwrite:
                    status_db, rid = upsert_result_by_start(conn, user_id, per_ini, per_fim, fname, float(total), df)
                else:
                    conn.execute("""
                        INSERT INTO results (user_id, created_at, period_start, period_end, file_name, total_esperado, items_json)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (user_id, datetime.utcnow().isoformat(), per_ini, per_fim, fname,
                          float(total), json.dumps(df.to_dict(orient="records"), ensure_ascii=False)))
                    conn.commit()
                    status_db, rid = ("inserted", conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            else:
                status_db, rid = ("simulated", 0)

            rows.append({
                "arquivo": fname, "period_start": per_ini, "period_end": per_fim,
                "total_esperado": float(total), "status": status_db
            })
            status.success(f"[{i}/{total_files}] OK: {fname} • {per_ini or '-'} — {per_fim or '-'} • {status_db}")
        except Exception as e:
            rows.append({
                "arquivo": fname, "period_start": None, "period_end": None,
                "total_esperado": 0.0, "status": f"erro: {e}"
            })
            status.error(f"[{i}/{total_files}] ERRO: {fname} • {e}")

        prog.progress(i / total_files)

    st.success("Lote finalizado.")
    df_sum = pd.DataFrame(rows, columns=["arquivo","period_start","period_end","total_esperado","status"])
    st.dataframe(df_sum, use_container_width=True, hide_index=True)

    # Exportar o resumo do lote
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine=("openpyxl" if _has_openpyxl() else "xlsxwriter")) as writer:
        df_sum.to_excel(writer, index=False, sheet_name="Resumo do lote")
    st.download_button(
        "Baixar resumo do lote (.xlsx)",
        buffer.getvalue(),
        "resumo_lote.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa
        return True
    except Exception:
        return False

def _fallback_upsert_by_start(conn, user_id: int, per_ini: Optional[str], per_fim: Optional[str],
                              file_name: str, total_esperado: float, df: pd.DataFrame) -> tuple[str,int]:
    """Caso você não tenha colado a upsert_result_by_start: atualiza se encontrar mesmo início; senão insere."""
    in_start_iso = _pdfdate_to_iso_str(per_ini)
    match_id = None
    cur = conn.execute("SELECT id, period_start FROM results WHERE user_id = ?", (user_id,))
    for rid, db_start in cur.fetchall():
        if _pdfdate_to_iso_str(db_start) == in_start_iso:
            match_id = rid
            break
    now_iso = datetime.utcnow().isoformat()
    payload = json.dumps(df.to_dict(orient="records"), ensure_ascii=False)

    if match_id is not None:
        conn.execute("""
            UPDATE results
               SET created_at=?, period_start=?, period_end=?, file_name=?, total_esperado=?, items_json=?
             WHERE id=? AND user_id=?
        """, (now_iso, per_ini, per_fim, file_name, total_esperado, payload, match_id, user_id))
        conn.commit()
        return ("updated", match_id)

    conn.execute("""
        INSERT INTO results (user_id, created_at, period_start, period_end, file_name, total_esperado, items_json)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, now_iso, per_ini, per_fim, file_name, total_esperado, payload))
    conn.commit()
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return ("inserted", rid)

# ============== HELPERS =====================
def _safe_rerun():
    """Roda o rerun em qualquer versão do Streamlit."""
    if hasattr(st, "rerun"):
        st.rerun()  # versões novas
    elif hasattr(st, "experimental_rerun"):
        st.experimental_rerun()  # versões antigas
    else:
        # fallback bruto
        st.session_state["_force_rerun"] = True
        st.warning("Recarregue a página (Ctrl+R).")



# =================== APP ===================
def main():
    conn = get_conn()
    auth_ui(conn)

    st.title("Wofair • Tarifação de Contas e Serviços")

    if "user" not in st.session_state:
        st.info("Entre ou crie uma conta para continuar.")
        st.stop()

    user = st.session_state["user"]
    st.success(f"Bem-vindo(a), {user['email']}!")
      
    #Função de deletar os dados
    if "user" in st.session_state:
        u = st.session_state["user"]

    with st.sidebar.expander("🧨 Zona de risco", expanded=False):
        st.caption("Ação irreversível!")
        choice = st.radio("O que apagar?", ["Só meus resultados", "Minha conta + tudo"], key="wipe_choice")
        confirm = st.text_input("Digite APAGAR para confirmar", key="wipe_confirm")
        go = st.button("Apagar agora")

        if go:
            if confirm.strip().upper() != "APAGAR":
                st.warning("Confirmação inválida. Digite exatamente APAGAR.")
            else:
                if choice == "Só meus resultados":
                    n = delete_results(get_conn(), u["id"])
                    st.success(f"Removidos {n} resultado(s).")
                else:
                    n_res, n_usr = delete_user_and_results(get_conn(), u["id"])
                    st.success(f"Conta apagada ({n_usr}) e resultados removidos ({n_res}).")
                    st.session_state.pop("user", None)
                    _safe_rerun()
                if st.button("Deduplicar resultados"):
                    n = dedupe_by_start(get_conn(), u["id"])
                    st.succes(f"Removidos {n} duplicados. Atualize a página!")

    tab_up, tab_hist, tab_rep = st.tabs(["📥 Uploads", "📜 Histórico", "📆 Relatório por Período"])
    
    with tab_up:
        render_uploads(conn, user)

    with tab_hist:
        render_historico(conn, user)

    with tab_rep:
        render_relatorio_periodo(conn, user)
        
        
       
    
        
#====================== CALL DO APP ======================
if __name__ == "__main__":
        main()